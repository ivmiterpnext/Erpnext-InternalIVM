"""
Read-only diagnostic: identify (item, bay) rows where the prod bin quantity
already matched the inventory spreadsheet before the reconciliation ran (and
were therefore skipped by inventory_recon.py's diff-based row selection), but
which have a non-zero price in the price sheet and a stale valuation_rate on
prod that was never updated.

Makes no changes. Run via:

    bench --site ivmportal.frappe.cloud execute "exec(open('/home/frappe/frappe-bench/sites/ivmportal.frappe.cloud/private/files/valuation_gap_diagnostic.py').read(), globals())"
"""

import frappe
import openpyxl
from frappe.utils import flt

SITE_FILES_DIR = "/home/frappe/frappe-bench/sites/ivmportal.frappe.cloud/private/files"
INVENTORY_FILE = f"{SITE_FILES_DIR}/Inventory 6.2026.xlsx"
PRICE_FILE = f"{SITE_FILES_DIR}/Parts list with Prices 7.2026.xlsx"

WAREHOUSE_MAP = {
    "Printshop-I": "Print Shop - I",
}

EXCLUDED_WAREHOUSES = {
    "Build In Progress - I",
    "Finished Goods - I",
    "Goods In Transit - I",
    "IT - I",
    "Stores - I",
    "Work In Progress - I",
    "3D Lab - I",
}


def log(msg):
    print(msg, flush=True)


def parse_inventory(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"]
    stock = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        bay, product_type, entry_name, sid, qty = row[0], row[1], row[2], row[3], row[4]
        if not bay or not str(bay).strip():
            continue
        if qty is None or str(qty).strip() == "":
            continue
        bay = str(bay).strip()
        bay = WAREHOUSE_MAP.get(bay, bay)
        sid = str(sid).strip()
        qty = flt(qty)
        key = (sid, bay)
        stock[key] = stock.get(key, 0.0) + qty
    return stock


def parse_prices(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"]
    prices = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        name, sid, price = row[0], row[1], row[2]
        if not sid or not str(sid).strip():
            continue
        sid = str(sid).strip()
        if sid in prices:
            continue
        if price is None or str(price).strip().upper() == "N/A":
            prices[sid] = 0.0
        else:
            prices[sid] = flt(price)
    return prices


def get_bay_warehouses():
    warehouses = frappe.get_all("Warehouse", filters={"is_group": 0}, pluck="name")
    return {w for w in warehouses if w not in EXCLUDED_WAREHOUSES}


def main():
    log("Parsing spreadsheets...")
    sheet_stock = parse_inventory(INVENTORY_FILE)
    prices = parse_prices(PRICE_FILE)
    bay_warehouses = get_bay_warehouses()

    item_codes = list({sid for sid, _ in sheet_stock})
    existing_items = set(
        frappe.get_all("Item", filters={"item_code": ["in", item_codes]}, pluck="item_code")
    )

    log("Loading current Bin state...")
    bins = frappe.get_all(
        "Bin",
        filters={"item_code": ["in", item_codes]},
        fields=["item_code", "warehouse", "actual_qty", "valuation_rate"],
    )
    bin_map = {(b.item_code, b.warehouse): b for b in bins}

    gap_rows = []
    for (item_code, bay), sheet_qty in sheet_stock.items():
        if item_code not in existing_items or bay not in bay_warehouses:
            continue
        price = prices.get(item_code, 0.0)
        if price <= 0:
            continue
        current = bin_map.get((item_code, bay))
        current_qty = flt(current.actual_qty) if current else 0.0
        current_rate = flt(current.valuation_rate) if current else 0.0
        if current_qty != sheet_qty:
            continue
        if abs(current_rate - price) > 0.001:
            gap_rows.append((item_code, bay, current_qty, current_rate, price))

    log(f"\nFound {len(gap_rows)} rows where qty already matched but valuation is stale:")
    for item_code, bay, qty, current_rate, price in sorted(gap_rows):
        log(f"  {item_code} @ {bay}: qty={qty}, current_rate={current_rate}, should_be={price}")

    log(f"\nTotal rows needing valuation-only correction: {len(gap_rows)}")


main()
