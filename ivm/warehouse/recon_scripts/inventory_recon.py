"""
Inventory reconciliation and price update script.

Run via bench execute inside the Frappe Cloud hybrid bench container:

    bench --site ivmportal.frappe.cloud execute "exec(open('/home/frappe/frappe-bench/sites/ivmportal.frappe.cloud/private/files/inventory_recon.py').read())"

DRY_RUN defaults to True. Review the printed report fully before setting DRY_RUN = False
and re-running. bench execute already initializes and connects frappe for the site;
this script must not call frappe.connect()/frappe.init()/frappe.destroy() itself.
"""

import frappe
import openpyxl
from frappe.utils import flt, nowdate

DRY_RUN = True

SITE_FILES_DIR = "/home/frappe/frappe-bench/sites/ivmportal.frappe.cloud/private/files"
INVENTORY_FILE = f"{SITE_FILES_DIR}/Inventory 6.2026.xlsx"
PRICE_FILE = f"{SITE_FILES_DIR}/Parts list with Prices 7.2026.xlsx"

PRICE_LIST = "Standard Buying"
CURRENCY = "USD"
COUNT_DATE = "2026-07-10"
BATCH_SIZE = 200

WAREHOUSE_MAP = {
    "Printshop-I": "Print Shop - I",
}

EXCLUDED_WAREHOUSES = {
    "Build In Progress - I",
    "Finished Goods - I",
    "Goods In Transit - I",
    "IT - I",
    "Stores - I",
    "Work In Progress - I",
    "3D Lab - I",
}


def log(msg):
    print(msg, flush=True)


def resolve_company():
    companies = frappe.get_all("Company", pluck="name")
    if len(companies) == 1:
        log(f"  Auto-detected single company: {companies[0]}")
        return companies[0]
    raise Exception(
        f"Expected exactly one Company on this site, found {len(companies)}: {companies}. "
        "Hardcode COMPANY at the top of this script and re-run."
    )


def parse_inventory(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"]
    stock = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        bay, product_type, entry_name, sid, qty = row[0], row[1], row[2], row[3], row[4]
        if not bay or not str(bay).strip():
            continue
        if qty is None or str(qty).strip() == "":
            continue
        bay = str(bay).strip()
        bay = WAREHOUSE_MAP.get(bay, bay)
        sid = str(sid).strip()
        qty = flt(qty)
        key = (sid, bay)
        stock[key] = stock.get(key, 0.0) + qty
    return stock


def parse_prices(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"]
    prices = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        name, sid, price = row[0], row[1], row[2]
        if not sid or not str(sid).strip():
            continue
        sid = str(sid).strip()
        if sid in prices:
            continue
        if price is None or str(price).strip().upper() == "N/A":
            prices[sid] = 0.0
        else:
            prices[sid] = flt(price)
    return prices


def get_bay_warehouses():
    warehouses = frappe.get_all("Warehouse", filters={"is_group": 0}, pluck="name")
    return {w for w in warehouses if w not in EXCLUDED_WAREHOUSES}


def validate_items(item_codes):
    item_codes = list(item_codes)
    existing = set(frappe.get_all("Item", filters={"item_code": ["in", item_codes]}, pluck="item_code"))
    missing = set(item_codes) - existing
    disabled = set(
        frappe.get_all(
            "Item",
            filters={"item_code": ["in", list(existing)], "disabled": 1},
            pluck="item_code",
        )
    )
    return existing, missing, disabled


def upsert_item_prices(prices, missing_items):
    created, updated = 0, 0
    for item_code, rate in prices.items():
        if item_code in missing_items:
            continue
        existing_name = frappe.db.get_value(
            "Item Price", {"item_code": item_code, "price_list": PRICE_LIST}, "name"
        )
        if existing_name:
            if not DRY_RUN:
                doc = frappe.get_doc("Item Price", existing_name)
                if doc.price_list_rate != rate:
                    doc.price_list_rate = rate
                    doc.currency = CURRENCY
                    doc.save()
            updated += 1
        else:
            if not DRY_RUN:
                doc = frappe.new_doc("Item Price")
                doc.item_code = item_code
                doc.price_list = PRICE_LIST
                doc.price_list_rate = rate
                doc.currency = CURRENCY
                doc.buying = 1
                doc.insert()
            created += 1
    return created, updated


def get_current_bay_stock(item_codes, bay_warehouses):
    if not item_codes:
        return {}
    rows = frappe.get_all(
        "Bin",
        filters={"item_code": ["in", list(item_codes)], "actual_qty": [">", 0]},
        fields=["item_code", "warehouse", "actual_qty"],
    )
    result = {}
    for r in rows:
        if r.warehouse not in bay_warehouses:
            continue
        result[(r.item_code, r.warehouse)] = flt(r.actual_qty)
    return result


def get_post_count_movements(item_codes, count_date):
    if not item_codes:
        return {}
    rows = frappe.db.sql(
        """
        SELECT item_code, warehouse, SUM(actual_qty) as net_qty
        FROM `tabStock Ledger Entry`
        WHERE item_code IN %(item_codes)s
        AND posting_date > %(count_date)s
        AND is_cancelled = 0
        GROUP BY item_code, warehouse
        """,
        {"item_codes": list(item_codes), "count_date": count_date},
        as_dict=True,
    )
    return {(r.item_code, r.warehouse): flt(r.net_qty) for r in rows}


def compute_reconciliation_rows(sheet_stock, current_bay_stock, movements, prices):
    rows = []
    warnings = []

    all_item_bay_pairs = set(sheet_stock.keys()) | set(current_bay_stock.keys())

    for item_code, bay in sorted(all_item_bay_pairs):
        in_sheet = (item_code, bay) in sheet_stock
        base_qty = sheet_stock.get((item_code, bay), 0.0) if in_sheet else 0.0
        net_movement = movements.get((item_code, bay), 0.0)
        adjusted_qty = base_qty - net_movement

        if adjusted_qty < 0:
            warnings.append(
                f"{item_code} @ {bay}: adjusted qty {adjusted_qty} < 0 "
                f"(base={base_qty}, movement={net_movement}) -- clamped to 0"
            )
            adjusted_qty = 0.0

        current_actual = current_bay_stock.get((item_code, bay), 0.0)
        if adjusted_qty == current_actual:
            continue

        rate = prices.get(item_code, 0.0)
        rows.append(
            {
                "item_code": item_code,
                "warehouse": bay,
                "qty": adjusted_qty,
                "valuation_rate": rate,
                "allow_zero_valuation_rate": 1 if rate == 0 else 0,
            }
        )

    return rows, warnings


def create_stock_reconciliations(rows, company):
    docs_created = []
    for i in range(0, len(rows), BATCH_SIZE):
        batch = rows[i : i + BATCH_SIZE]
        if DRY_RUN:
            log(f"  [DRY RUN] Would create Stock Reconciliation with {len(batch)} rows")
            continue
        doc = frappe.new_doc("Stock Reconciliation")
        doc.company = company
        doc.purpose = "Stock Reconciliation"
        doc.posting_date = nowdate()
        for row in batch:
            doc.append("items", row)
        doc.insert()
        doc.submit()
        docs_created.append(doc.name)
        log(f"  Created and submitted: {doc.name} ({len(batch)} rows)")
    return docs_created


def main():
    log("=" * 70)
    log(f"INVENTORY RECONCILIATION SCRIPT -- DRY_RUN={DRY_RUN}")
    log("=" * 70)

    log("\nResolving company...")
    company = resolve_company()

    log("\nParsing inventory spreadsheet...")
    sheet_stock = parse_inventory(INVENTORY_FILE)
    log(f"  Parsed {len(sheet_stock)} (item, bay) entries")

    log("\nParsing price spreadsheet...")
    prices = parse_prices(PRICE_FILE)
    log(f"  Parsed {len(prices)} unique item prices")

    all_item_codes = {sid for sid, _ in sheet_stock} | set(prices.keys())
    all_warehouse_names = {bay for _, bay in sheet_stock}

    log(f"\nValidating {len(all_item_codes)} items against site...")
    existing_items, missing_items, disabled_items = validate_items(all_item_codes)
    if missing_items:
        log(f"  WARNING: {len(missing_items)} items not found on site:")
        for i in sorted(missing_items):
            log(f"    - {i}")
    if disabled_items:
        log(f"  WARNING: {len(disabled_items)} items are disabled:")
        for i in sorted(disabled_items):
            log(f"    - {i}")

    log("\nLoading bay-type warehouses...")
    bay_warehouses = get_bay_warehouses()
    log(f"  Found {len(bay_warehouses)} bay-type warehouses on site")

    missing_warehouses = {w for w in all_warehouse_names if w not in bay_warehouses}
    if missing_warehouses:
        log(f"  WARNING: {len(missing_warehouses)} warehouses from spreadsheet not found as bay-type warehouses:")
        for w in sorted(missing_warehouses):
            log(f"    - {w}")

    log("\nFiltering out missing items/warehouses from reconciliation set...")
    sheet_stock = {
        (sid, bay): qty
        for (sid, bay), qty in sheet_stock.items()
        if sid in existing_items and sid not in disabled_items and bay not in missing_warehouses
    }
    valid_item_codes = {sid for sid, _ in sheet_stock}
    log(f"  {len(sheet_stock)} valid (item, bay) entries remain")

    log(f"\nUpserting Item Price records ({PRICE_LIST})...")
    created, updated = upsert_item_prices(prices, missing_items)
    log(f"  Created: {created}, Updated/checked: {updated}")

    log(f"\nQuerying current bay stock on site for {len(valid_item_codes)} items...")
    current_bay_stock = get_current_bay_stock(valid_item_codes, bay_warehouses)
    log(f"  Found {len(current_bay_stock)} (item, bay) entries with existing stock")

    log(f"\nQuerying post-count-date ({COUNT_DATE}) stock movements...")
    movements = get_post_count_movements(valid_item_codes, COUNT_DATE)
    log(f"  Found movement records for {len(movements)} (item, bay) combinations")

    log("\nComputing adjusted reconciliation targets...")
    rows, warnings = compute_reconciliation_rows(sheet_stock, current_bay_stock, movements, prices)
    log(f"  {len(rows)} rows require reconciliation (differ from current state)")
    if warnings:
        log(f"  {len(warnings)} warnings:")
        for w in warnings[:50]:
            log(f"    - {w}")
        if len(warnings) > 50:
            log(f"    ... and {len(warnings) - 50} more")

    total_value = sum(r["qty"] * r["valuation_rate"] for r in rows)
    log(f"\n  Total reconciliation value: ${total_value:,.2f}")

    log(f"\nCreating Stock Reconciliation document(s) in batches of {BATCH_SIZE}...")
    docs = create_stock_reconciliations(rows, company)

    log("\n" + "=" * 70)
    log("SUMMARY")
    log("=" * 70)
    log(f"  Company: {company}")
    log(f"  Items processed: {len(valid_item_codes)}")
    log(f"  Missing items (skipped): {len(missing_items)}")
    log(f"  Disabled items (skipped): {len(disabled_items)}")
    log(f"  Missing warehouses (skipped): {len(missing_warehouses)}")
    log(f"  Item Prices created: {created}")
    log(f"  Item Prices updated/checked: {updated}")
    log(f"  Reconciliation rows: {len(rows)}")
    log(f"  Warnings: {len(warnings)}")
    if not DRY_RUN:
        log(f"  Stock Reconciliation docs submitted: {len(docs)}")
        for d in docs:
            log(f"    - {d}")
    else:
        log("  DRY RUN -- no documents were created. Set DRY_RUN=False to execute.")
    log("=" * 70)


main()

if not DRY_RUN:
    frappe.db.commit()
    log("\nChanges committed.")
else:
    frappe.db.rollback()
