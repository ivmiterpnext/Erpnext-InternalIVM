"""
Valuation-only remediation script.

Two independent fixes:

1. Item-level: directly sets tabItem.valuation_rate for every item with a
   non-zero price in the price sheet. This field is a plain, directly
   editable Currency field on Item, NOT read-only, NOT tied to Bin or the
   Stock Ledger -- it is a single global reference value per item,
   separate entirely from per-warehouse stock costing. Safe to update with
   a direct frappe.db.set_value call, no Stock Reconciliation needed.

2. Bin-level: fixes (item, bay) rows where the prod bin quantity already
   matched the inventory spreadsheet (so the original inventory_recon.py
   run correctly skipped changing quantity for them) AND current qty > 0,
   which have a non-zero price in the price sheet and a stale
   valuation_rate that was never applied. Each affected row is reconciled
   with qty UNCHANGED (set to its current actual_qty) and valuation_rate
   corrected to the price sheet value -- a standard ERPNext "revaluation"
   pattern where Stock Reconciliation sets the same qty with a different
   rate, which creates a new Stock Ledger Entry.

Rows where current qty == 0 are deliberately EXCLUDED from the Bin-level
fix: ERPNext's Stock Reconciliation.update_stock_ledger() skips creating
a Stock Ledger Entry whenever a row's qty stays at zero (see the
`row.qty == 0` clause in stock_reconciliation.py's continue condition),
regardless of valuation_rate, so attempting it would silently no-op. Since
inventory value = qty * rate, a zero-qty bin's stale rate has zero actual
financial impact -- 0 times anything is 0 -- so this is intentionally left
alone rather than forcing a synthetic qty-up-then-down transaction pair
just to cosmetically update a read-only, zero-impact field. The item-level
fix above still gives these items a correct reference valuation_rate.

BATCH_SIZE is deliberately capped at 100. Stock Reconciliation.submit()
defers to a background job for any document with MORE than 100 rows
(len(self.items) > 100), which caused a MySQL deadlock across concurrently
enqueued background jobs during the original run when BATCH_SIZE=200 was
used. Keeping batches at exactly 100 keeps every submit() call synchronous
and in-process, avoiding that failure mode entirely.

DRY_RUN defaults to True. Review the printed report fully before setting
DRY_RUN = False and re-running. bench execute already initializes and
connects frappe for the site; this script must not call
frappe.connect()/frappe.init()/frappe.destroy() itself.

Run via:

    bench --site ivmportal.frappe.cloud execute "exec(open('/home/frappe/frappe-bench/sites/ivmportal.frappe.cloud/private/files/valuation_remediation.py').read(), globals())"
"""

import frappe
import openpyxl
from frappe.utils import flt, nowdate

DRY_RUN = True

SITE_FILES_DIR = "/home/frappe/frappe-bench/sites/ivmportal.frappe.cloud/private/files"
INVENTORY_FILE = f"{SITE_FILES_DIR}/Inventory 6.2026.xlsx"
PRICE_FILE = f"{SITE_FILES_DIR}/Parts list with Prices 7.2026.xlsx"

BATCH_SIZE = 100

WAREHOUSE_MAP = {
    "Printshop-I": "Print Shop - I",
}

EXCLUDED_WAREHOUSES = {
    "Build In Progress - I",
    "Finished Goods - I",
    "Goods In Transit - I",
    "IT - I",
    "Stores - I",
    "Work In Progress - I",
    "3D Lab - I",
}


def log(msg):
    print(msg, flush=True)


def resolve_company():
    companies = frappe.get_all("Company", pluck="name")
    if len(companies) == 1:
        log(f"  Auto-detected single company: {companies[0]}")
        return companies[0]
    raise Exception(
        f"Expected exactly one Company on this site, found {len(companies)}: {companies}. "
        "Hardcode COMPANY at the top of this script and re-run."
    )


def parse_inventory(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"]
    stock = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        bay, product_type, entry_name, sid, qty = row[0], row[1], row[2], row[3], row[4]
        if not bay or not str(bay).strip():
            continue
        if qty is None or str(qty).strip() == "":
            continue
        bay = str(bay).strip()
        bay = WAREHOUSE_MAP.get(bay, bay)
        sid = str(sid).strip()
        qty = flt(qty)
        key = (sid, bay)
        stock[key] = stock.get(key, 0.0) + qty
    return stock


def parse_prices(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"]
    prices = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        name, sid, price = row[0], row[1], row[2]
        if not sid or not str(sid).strip():
            continue
        sid = str(sid).strip()
        if sid in prices:
            continue
        if price is None or str(price).strip().upper() == "N/A":
            prices[sid] = 0.0
        else:
            prices[sid] = flt(price)
    return prices


def get_bay_warehouses():
    warehouses = frappe.get_all("Warehouse", filters={"is_group": 0}, pluck="name")
    return {w for w in warehouses if w not in EXCLUDED_WAREHOUSES}


def update_item_valuation_rates(prices):
    updated, skipped_no_price, already_correct = 0, 0, 0
    for item_code, price in prices.items():
        if price <= 0:
            skipped_no_price += 1
            continue
        current = frappe.db.get_value("Item", item_code, "valuation_rate")
        if current is None:
            skipped_no_price += 1
            continue
        if abs(flt(current) - price) <= 0.001:
            already_correct += 1
            continue
        if not DRY_RUN:
            frappe.db.set_value("Item", item_code, "valuation_rate", price, update_modified=False)
        updated += 1
    return updated, skipped_no_price, already_correct


def find_gap_rows(sheet_stock, prices, bay_warehouses):
    item_codes = list({sid for sid, _ in sheet_stock})
    existing_items = set(
        frappe.get_all("Item", filters={"item_code": ["in", item_codes]}, pluck="item_code")
    )

    bins = frappe.get_all(
        "Bin",
        filters={"item_code": ["in", item_codes]},
        fields=["item_code", "warehouse", "actual_qty", "valuation_rate"],
    )
    bin_map = {(b.item_code, b.warehouse): b for b in bins}

    gap_rows = []
    for (item_code, bay), sheet_qty in sheet_stock.items():
        if item_code not in existing_items or bay not in bay_warehouses:
            continue
        price = prices.get(item_code, 0.0)
        if price <= 0:
            continue
        current = bin_map.get((item_code, bay))
        current_qty = flt(current.actual_qty) if current else 0.0
        current_rate = flt(current.valuation_rate) if current else 0.0
        if current_qty != sheet_qty:
            continue
        if current_qty == 0:
            continue
        if abs(current_rate - price) > 0.001:
            gap_rows.append(
                {
                    "item_code": item_code,
                    "warehouse": bay,
                    "qty": current_qty,
                    "valuation_rate": price,
                    "allow_zero_valuation_rate": 0,
                }
            )
    return gap_rows


def create_stock_reconciliations(rows, company):
    docs_created = []
    for i in range(0, len(rows), BATCH_SIZE):
        batch = rows[i : i + BATCH_SIZE]
        if DRY_RUN:
            log(f"  [DRY RUN] Would create Stock Reconciliation with {len(batch)} rows")
            continue
        doc = frappe.new_doc("Stock Reconciliation")
        doc.company = company
        doc.purpose = "Stock Reconciliation"
        doc.posting_date = nowdate()
        for row in batch:
            doc.append("items", row)
        doc.insert()
        doc.submit()
        docs_created.append(doc.name)
        log(f"  Created and submitted: {doc.name} ({len(batch)} rows)")
    return docs_created


def main():
    log("=" * 70)
    log(f"VALUATION-ONLY REMEDIATION SCRIPT -- DRY_RUN={DRY_RUN}")
    log("=" * 70)

    log("\nResolving company...")
    company = resolve_company()

    log("\nParsing inventory spreadsheet...")
    sheet_stock = parse_inventory(INVENTORY_FILE)
    log(f"  Parsed {len(sheet_stock)} (item, bay) entries")

    log("\nParsing price spreadsheet...")
    prices = parse_prices(PRICE_FILE)
    log(f"  Parsed {len(prices)} unique item prices")

    log("\nLoading bay-type warehouses...")
    bay_warehouses = get_bay_warehouses()
    log(f"  Found {len(bay_warehouses)} bay-type warehouses on site")

    log("\nUpdating Item.valuation_rate directly (item-level, no Bin/SLE interaction)...")
    item_updated, item_skipped_no_price, item_already_correct = update_item_valuation_rates(prices)
    log(f"  Updated: {item_updated}, Already correct: {item_already_correct}, Skipped (no price): {item_skipped_no_price}")

    log("\nFinding rows with matching qty but stale valuation...")
    rows = find_gap_rows(sheet_stock, prices, bay_warehouses)
    log(f"  Found {len(rows)} rows needing valuation-only correction")

    total_value_shift = sum(r["qty"] * r["valuation_rate"] for r in rows)
    log(f"\n  Total value being applied via this correction: ${total_value_shift:,.2f}")

    if len(rows) <= 20:
        log("\n  All affected rows:")
        for r in sorted(rows, key=lambda x: (x["item_code"], x["warehouse"])):
            log(f"    {r['item_code']} @ {r['warehouse']}: qty={r['qty']} (unchanged), rate -> {r['valuation_rate']}")
    else:
        log("\n  First 20 affected rows:")
        for r in sorted(rows, key=lambda x: (x["item_code"], x["warehouse"]))[:20]:
            log(f"    {r['item_code']} @ {r['warehouse']}: qty={r['qty']} (unchanged), rate -> {r['valuation_rate']}")
        log(f"    ... and {len(rows) - 20} more")

    log(f"\nCreating Stock Reconciliation document(s) in batches of {BATCH_SIZE}...")
    docs = create_stock_reconciliations(rows, company)

    log("\n" + "=" * 70)
    log("SUMMARY")
    log("=" * 70)
    log(f"  Company: {company}")
    log(f"  Item.valuation_rate updated: {item_updated}")
    log(f"  Item.valuation_rate already correct: {item_already_correct}")
    log(f"  Rows corrected: {len(rows)}")
    log(f"  Total value applied: ${total_value_shift:,.2f}")
    if not DRY_RUN:
        log(f"  Stock Reconciliation docs submitted: {len(docs)}")
        for d in docs:
            log(f"    - {d}")
    else:
        log("  DRY RUN -- no documents were created. Set DRY_RUN=False to execute.")
    log("=" * 70)


main()

if not DRY_RUN:
    frappe.db.commit()
    log("\nChanges committed.")
else:
    frappe.db.rollback()
